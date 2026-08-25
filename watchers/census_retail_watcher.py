import io
import os
import re
import time
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .common import (
    BaseWatcher,
    Event,
    REQUEST_TIMEOUT,
    USER_AGENT,
    WEB_POLL_SECONDS,
    WARMUP_PREVIEW_COUNT,
    apply_warmup_mode,
    bounded_timeout_seconds,
    clean_text,
    fetch_with_retries,
    make_instance_source_name,
    parse_csv_env,
    stable_id,
)

DEFAULT_CENSUS_RETAIL_URL = "https://www.census.gov/retail/sales.html"
DEFAULT_CENSUS_RETAIL_PDF_PATH = "marts/www/marts_current.pdf"
DEFAULT_CENSUS_RETAIL_XLSX_PATH = "marts/www/marts_current.xlsx"
CENSUS_RETAIL_DEFAULT_PAGE_5XX_COOLDOWN_SECONDS = 1800.0
RELEASE_RE = re.compile(r"FOR IMMEDIATE RELEASE:\s*(.+)$", re.IGNORECASE)
SUMMARY_RE = re.compile(
    r"sales for (?P<period>[A-Za-z]+\s+\d{4}).+?were \$(?P<sales>[\d.]+) billion, "
    r"up (?P<pct>[-\d.]+) percent",
    re.IGNORECASE,
)
MONTH_NAMES = {
    "jan": "January",
    "feb": "February",
    "mar": "March",
    "apr": "April",
    "may": "May",
    "jun": "June",
    "jul": "July",
    "aug": "August",
    "sep": "September",
    "oct": "October",
    "nov": "November",
    "dec": "December",
}


class CensusRetailWatcher(BaseWatcher):
    def __init__(self, source_name: str, url: str, interval_seconds: int = WEB_POLL_SECONDS):
        super().__init__(source_name, interval_seconds)
        self.url = url
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.page_5xx_cooldown_until = 0.0
        self.page_5xx_cooldown_seconds = max(
            60.0,
            float(
                os.getenv(
                    "CENSUS_RETAIL_PAGE_5XX_COOLDOWN_SECONDS",
                    str(CENSUS_RETAIL_DEFAULT_PAGE_5XX_COOLDOWN_SECONDS),
                )
                or CENSUS_RETAIL_DEFAULT_PAGE_5XX_COOLDOWN_SECONDS
            ),
        )

    def _fetch_page_html(self) -> str:
        resp = self.session.get(
            self.url,
            timeout=bounded_timeout_seconds(REQUEST_TIMEOUT),
            headers={"Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.1"},
        )
        resp.raise_for_status()
        return resp.text

    def _fetch_binary(self, url: str) -> bytes:
        resp = fetch_with_retries(self.session, url)
        return resp.content

    def _current_report_url(self, path: str) -> str:
        return urljoin(self.url, path)

    def _extract_page_metadata(self, html: str) -> dict[str, str]:
        soup = BeautifulSoup(html, "html.parser")
        title = ""
        release_date = ""
        summary = ""
        pdf_url = ""
        xlsx_url = ""

        for text in soup.stripped_strings:
            cleaned = clean_text(text)
            if not cleaned:
                continue
            if not release_date:
                match = RELEASE_RE.search(cleaned)
                if match:
                    release_date = clean_text(match.group(1))
            if not title and cleaned.lower() == "advance monthly sales for retail and food services":
                title = cleaned
            if not summary and cleaned.lower().startswith("advance estimates of u.s. retail and food services sales for"):
                summary = cleaned

        for anchor in soup.find_all("a", href=True):
            text = clean_text(anchor.get_text(" ", strip=True))
            href = urljoin(self.url, anchor["href"])
            if not pdf_url and text == "Advance Monthly Retail Trade Report":
                pdf_url = href
            elif not xlsx_url and text == "Advance Monthly Retail Trade Report Tables":
                xlsx_url = href

        if not title or not summary or not xlsx_url:
            raise RuntimeError("could not parse Census retail-sales release metadata")

        return {
            "title": title,
            "release_date": release_date,
            "summary": summary,
            "pdf_url": pdf_url or self.url,
            "xlsx_url": xlsx_url,
        }

    def _find_adjusted_start_column(self, ws) -> int:
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for index, value in enumerate(row, start=1):
                if clean_text(str(value or "")).lower().startswith("adjusted"):
                    return index
        raise RuntimeError("could not find adjusted data columns in Census workbook")

    def _extract_report_period(self, ws, adjusted_start_col: int) -> str:
        year = ""
        month = ""
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            value = row[adjusted_start_col - 1] if len(row) >= adjusted_start_col else None
            if isinstance(value, (int, float)) and 2000 <= int(value) <= 2100:
                year = str(int(value))
                continue
            token = re.sub(r"[^A-Za-z]+", "", str(value or "")).lower()
            if len(token) >= 3:
                month = MONTH_NAMES.get(token[:3], month)
        if month and year:
            return f"{month} {year}"
        return ""

    def _extract_numeric_row_values(self, row: tuple, adjusted_start_col: int) -> tuple[float, float]:
        current = row[adjusted_start_col - 1]
        previous = row[adjusted_start_col]
        if not isinstance(current, (int, float)) or not isinstance(previous, (int, float)):
            raise RuntimeError("unexpected Census workbook row shape")
        return float(current), float(previous)

    def _parse_current_metrics(self, workbook_bytes: bytes) -> dict[str, Any]:
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        adjusted_start_col = self._find_adjusted_start_column(ws)
        report_period = self._extract_report_period(ws, adjusted_start_col)
        total_values = None
        ex_auto_values = None
        label_prefix = ""
        for row in ws.iter_rows(values_only=True):
            label = clean_text(str((row[1] or "") if len(row) > 1 else ""))
            if not label:
                continue
            numeric_window = row[adjusted_start_col - 1 : adjusted_start_col + 1]
            has_current_values = all(isinstance(item, (int, float)) for item in numeric_window)
            if not has_current_values:
                label_prefix = label
                continue
            combined_label = clean_text(f"{label_prefix} {label}")
            normalized = re.sub(r"[^a-z0-9]+", " ", combined_label.lower()).strip()
            if not total_values and normalized.startswith("retail food services total"):
                total_values = self._extract_numeric_row_values(row, adjusted_start_col)
            elif not ex_auto_values and "total excl motor vehicle parts" in normalized:
                ex_auto_values = self._extract_numeric_row_values(row, adjusted_start_col)
            if total_values and ex_auto_values:
                break

        if not total_values or not ex_auto_values:
            raise RuntimeError("could not locate current Census retail-sales rows in workbook")

        total_current, total_previous = total_values
        ex_auto_current, ex_auto_previous = ex_auto_values
        return {
            "total_sales_bil": total_current / 1000.0,
            "total_pct": ((total_current - total_previous) / total_previous) * 100.0,
            "ex_auto_sales_bil": ex_auto_current / 1000.0,
            "ex_auto_pct": ((ex_auto_current - ex_auto_previous) / ex_auto_previous) * 100.0,
            "report_period": report_period,
        }

    def _build_current_event_from_parts(
        self,
        *,
        title_base: str,
        release_date: str,
        summary_base: str,
        pdf_url: str,
        xlsx_url: str,
        page_fallback_reason: str = "",
    ) -> Event:
        metrics = self._parse_current_metrics(self._fetch_binary(xlsx_url))
        summary_match = SUMMARY_RE.search(summary_base)
        report_period = summary_match.group("period") if summary_match else str(metrics.get("report_period") or "")
        title = (
            f"{title_base}: {metrics['total_pct']:.1f}% m/m, "
            f"ex-auto {metrics['ex_auto_pct']:.1f}% m/m ({report_period})"
        )
        if summary_base:
            summary_prefix = summary_base
        else:
            summary_prefix = (
                f"Advance estimates of U.S. retail and food services sales for {report_period} "
                f"were ${metrics['total_sales_bil']:.1f} billion, "
                f"up {metrics['total_pct']:.1f} percent from the previous month."
            )
        summary = (
            f"{summary_prefix} Official workbook totals: adjusted sales ${metrics['total_sales_bil']:.1f}B; "
            f"ex-auto ${metrics['ex_auto_sales_bil']:.1f}B."
        )
        item_key = pdf_url if pdf_url != self.url else f"{report_period}|{metrics['total_pct']:.1f}|{metrics['ex_auto_pct']:.1f}"
        raw = {
            "page_url": self.url,
            "xlsx_url": xlsx_url,
            "report_period": report_period,
            "total_sales_bil": round(metrics["total_sales_bil"], 1),
            "total_pct": round(metrics["total_pct"], 1),
            "ex_auto_sales_bil": round(metrics["ex_auto_sales_bil"], 1),
            "ex_auto_pct": round(metrics["ex_auto_pct"], 1),
        }
        if page_fallback_reason:
            raw["page_fallback_reason"] = page_fallback_reason
        return Event(
            source=self.source_name,
            item_id=stable_id(self.source_name, item_key),
            title=title,
            url=pdf_url,
            published_at=release_date or report_period or None,
            summary=summary,
            category="retail_sales",
            raw=raw,
        )

    def _build_current_event(self, html: str) -> Event:
        meta = self._extract_page_metadata(html)
        return self._build_current_event_from_parts(
            title_base=meta["title"],
            release_date=meta["release_date"],
            summary_base=meta["summary"],
            pdf_url=meta["pdf_url"],
            xlsx_url=meta["xlsx_url"],
        )

    def _build_current_event_from_fallback_files(self, reason: str) -> Event:
        return self._build_current_event_from_parts(
            title_base="Advance Monthly Sales for Retail and Food Services",
            release_date="",
            summary_base="",
            pdf_url=self._current_report_url(DEFAULT_CENSUS_RETAIL_PDF_PATH),
            xlsx_url=self._current_report_url(DEFAULT_CENSUS_RETAIL_XLSX_PATH),
            page_fallback_reason=reason,
        )

    def _fetch_events(self) -> list[Event]:
        if time.time() < self.page_5xx_cooldown_until:
            return [self._build_current_event_from_fallback_files("page_5xx_cooldown")]
        try:
            return [self._build_current_event(self._fetch_page_html())]
        except requests.HTTPError as exc:
            status_code = getattr(getattr(exc, "response", None), "status_code", None)
            if status_code in {500, 502, 503, 504}:
                self.page_5xx_cooldown_until = max(
                    self.page_5xx_cooldown_until,
                    time.time() + self.page_5xx_cooldown_seconds,
                )
                return [self._build_current_event_from_fallback_files(f"page_http_{status_code}")]
            raise

    def warmup(self, state) -> list[Event]:
        items = self._fetch_events()
        preview = items[:WARMUP_PREVIEW_COUNT]
        print(f"[warmup][{self.source_name}] latest {len(preview)} items:")
        for item in preview:
            print(f"  title={item.title[:140]}")
            print(f"    url={item.url}")
        return apply_warmup_mode(
            self.source_name,
            state,
            [item.item_id for item in items],
            emitted_events=list(reversed(items)),
        )

    def poll(self, state) -> list[Event]:
        items = self._fetch_events()
        self.last_poll_at = time.time()
        new_items = [item for item in items if not state.is_seen(self.source_name, item.item_id)]
        for item in new_items:
            state.mark_seen(self.source_name, item.item_id)
        return list(reversed(new_items))


def build_census_retail_watchers() -> list[CensusRetailWatcher]:
    urls = parse_csv_env("CENSUS_RETAIL_URLS", DEFAULT_CENSUS_RETAIL_URL)
    total = len(urls)
    interval = int(os.getenv("CENSUS_RETAIL_POLL_SECONDS", str(WEB_POLL_SECONDS)))
    return [CensusRetailWatcher(make_instance_source_name("census_retail", i, total), url, interval) for i, url in enumerate(urls)]
