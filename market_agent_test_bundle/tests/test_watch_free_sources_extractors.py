from __future__ import annotations

import sys
import time
import xml.etree.ElementTree as ET
from types import SimpleNamespace
from io import BytesIO
from pathlib import Path

import pytest
import requests
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import watchers.aaa_watcher as aaa_module
from watchers.aaa_watcher import AAAWatcher
from watchers.adp_watcher import AdpWatcher
import watchers.axios_watcher as axios_module
from watchers.axios_watcher import AxiosWatcher, SitemapEntry
from watchers.bls_watcher import BlsWatcher
import watchers.bloomberg_watcher as bloomberg_module
from watchers.bloomberg_watcher import BloombergWatcher, _keywords_for_target
from watchers.census_retail_watcher import CensusRetailWatcher
from watchers.coindesk_watcher import CoinDeskWatcher
from watchers.conference_board_watcher import ConferenceBoardWatcher
from watchers.schwab_watcher import SchwabWatcher
from watchers.spglobal_pmi_watcher import SpglobalPmiWatcher
from watchers.dol_watcher import DolWatcher
from watchers.factbase_watcher import FactbaseWatcher
from watchers.irna_watcher import IrnaWatcher
from watchers.mni_watcher import MniWatcher
from watchers.nyt_watcher import NytFeedEntry, NytWatcher
from watchers.common import Event, HtmlPageWatcher, StateStore, WEB_POLL_SECONDS, fetch_with_retries
import watchers.reuters_watcher as reuters_module
from watchers.reuters_watcher import ReutersWatcher
from watchers.white_house_watcher import WhiteHouseWatcher
from watchers.treasury_watcher import TreasuryWatcher
from watchers.wsj_watcher import WsjFeedEntry, WsjWatcher

def test_aaa_watcher_skips_tools_and_nav_links():
    watcher = AAAWatcher("aaa_test", "https://gasprices.aaa.com/")
    html = """
    <html><body>
      <a href="/ev-charging-prices/">EV Charging Prices</a>
      <a href="/aaa-gas-cost-calculator/">Gas Cost Calculator</a>
      <a href="/national-gas-average-jumps-one-dollar-in-one-month/">National Gas Average Jumps One Dollar in One Month</a>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert [item.url for item in items] == [
        "https://gasprices.aaa.com/national-gas-average-jumps-one-dollar-in-one-month/"
    ]


def test_aaa_watcher_uses_browser_fallback_after_403(monkeypatch):
    watcher = AAAWatcher("aaa_test", "https://gasprices.aaa.com/")

    response = requests.Response()
    response.status_code = 403
    response.url = watcher.url
    http_error = requests.HTTPError("403 Client Error: Forbidden", response=response)

    def fail_fetch(session, url, *, accept=None, proxy_failover_statuses=None):
        raise http_error

    browser_calls = []

    class BrowserResponse:
        text = "<html><body>ok</body></html>"

        def raise_for_status(self):
            return None

    def browser_fetch(url, *, headers=None, accept=None):
        browser_calls.append((url, accept))
        return BrowserResponse()

    monkeypatch.setattr(aaa_module, "fetch_with_retries", fail_fetch)
    monkeypatch.setattr(aaa_module, "browser_impersonated_get", browser_fetch)

    assert watcher._fetch_html() == "<html><body>ok</body></html>"
    assert browser_calls == [(watcher.url, "text/html,application/xhtml+xml;q=0.9,*/*;q=0.1")]


def test_fetch_with_retries_rotates_source_proxy_after_403(monkeypatch):
    session = requests.Session()
    rotations: list[str] = []
    session._source_name = "aaa"
    session._source_proxy_failover_max_rotations = 1
    session._source_proxy_failover_sleep_seconds = 0
    session._source_proxy_failover = type(
        "DummyFailover",
        (),
        {
            "selector_group": "AAA Proxy",
            "rotate_to_next_node": lambda self: rotations.append("node-b") or "node-b",
        },
    )()
    calls = {"count": 0}

    def fake_get(url, timeout=None, headers=None):
        calls["count"] += 1
        response = requests.Response()
        response.url = url
        response.status_code = 403 if calls["count"] == 1 else 200
        response._content = b"ok"
        return response

    monkeypatch.setattr(session, "get", fake_get)

    response = fetch_with_retries(session, "https://gasprices.aaa.com/")

    assert response.status_code == 200
    assert calls["count"] == 2
    assert rotations == ["node-b"]


def test_aaa_watcher_skips_browser_fallback_without_403(monkeypatch):
    watcher = AAAWatcher("aaa_test", "https://gasprices.aaa.com/")

    class Response:
        text = "<html><body>plain</body></html>"

    browser_calls = []

    def ok_fetch(session, url, *, accept=None, proxy_failover_statuses=None):
        return Response()

    def browser_fetch(url, *, headers=None, accept=None):
        browser_calls.append((url, accept))
        raise AssertionError("browser fallback should not be used")

    monkeypatch.setattr(aaa_module, "fetch_with_retries", ok_fetch)
    monkeypatch.setattr(aaa_module, "browser_impersonated_get", browser_fetch)

    assert watcher._fetch_html() == "<html><body>plain</body></html>"
    assert browser_calls == []


def test_adp_watcher_parses_current_report_and_press_release(monkeypatch):
    watcher = AdpWatcher("adp_test", "https://adpemploymentreport.com/")
    html = """
    <html><body>
      <h1>ADP National Employment Report</h1>
      <div>February 2026</div>
      <div>up 63,000</div>
      <h2>Private employers added 63,000 jobs in February</h2>
      <p>Hiring jumped in February, delivering the best showing for job gains since November 2025.</p>
      <a href="/files/adp-feb-2026.pdf">View press release</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)

    items = watcher._fetch_events()

    assert len(items) == 1
    assert items[0].title == "Private employers added 63,000 jobs in February"
    assert items[0].url == "https://adpemploymentreport.com/files/adp-feb-2026.pdf"
    assert items[0].raw["report_month"] == "February 2026"
    assert items[0].raw["change_value"] == "up 63,000"


def test_conference_board_watcher_skips_media_contact_page():
    watcher = ConferenceBoardWatcher("conference_board_test", "https://www.conference-board.org/press/")
    html = """
    <html><body>
      <a href="/press/media-contact">media contacts and experts page</a>
      <a href="/press/where-to-hire-index">The Best and Worst Places to Hire in Europe January 12, 2026</a>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert len(items) == 1
    assert items[0].url == "https://www.conference-board.org/press/where-to-hire-index"


def test_spglobal_pmi_watcher_extracts_research_analysis_links():
    watcher = SpglobalPmiWatcher("spglobal_pmi_test", "https://www.spglobal.com/marketintelligence/en/mi/research-analysis/pmi.html")
    html = """
    <html><body>
      <a href="/marketintelligence/en/mi/research-analysis/pmi.html">PMI home</a>
      <a href="/marketintelligence/en/mi/research-analysis/week-ahead-economic-preview-week-of-2-march-2026.html">Week Ahead Economic Preview: Week of 2 March 2026</a>
      <a href="/marketintelligence/en/mi/research-analysis/monthly-pmi-bulletin-february-2026.html">Monthly PMI Bulletin: February 2026</a>
      <a href="/marketintelligence/en/mi/products/pmi.html">PMI Product Page</a>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert [item.title for item in items] == [
        "Week Ahead Economic Preview: Week of 2 March 2026",
        "Monthly PMI Bulletin: February 2026",
    ]


def test_schwab_watcher_extracts_story_cards_and_metadata():
    watcher = SchwabWatcher("schwab_test", "https://www.schwab.com/learn/market-commentary")
    html = """
    <html><body>
      <a href="/learn/topic/stocks">Stocks</a>
      <a href="/learn/story/what-iran-conflict-means-stocks-bonds-and-inflation" aria-label="What the Iran Conflict Could Mean for Stocks, Bonds &amp; Inflation">
        <h3>What the Iran Conflict Could Mean for Stocks, Bonds &amp; Inflation</h3>
        <span>What are some of the economic ripple effects stemming from the war in Iran?</span>
        <span>Podcast | Apr 3, 2026</span>
      </a>
      <a href="/learn/story/what-iran-conflict-means-stocks-bonds-and-inflation">
        <h3>What the Iran Conflict Could Mean for Stocks, Bonds &amp; Inflation</h3>
      </a>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert len(items) == 1
    assert items[0].url == "https://www.schwab.com/learn/story/what-iran-conflict-means-stocks-bonds-and-inflation"
    assert items[0].title == "What the Iran Conflict Could Mean for Stocks, Bonds & Inflation"
    assert items[0].summary == "What are some of the economic ripple effects stemming from the war in Iran?"
    assert items[0].published_at == "2026-04-03T00:00:00+00:00"


def test_treasury_watcher_skips_collection_and_nav_pages():
    watcher = TreasuryWatcher("treasury_test", "https://home.treasury.gov/news/press-releases")
    html = """
    <html><body>
      <a href="/news/press-releases/sb0429">Secretary Bessent: “Financial Literacy Unlocks Opportunity for Every American”</a>
      <a href="/news/press-releases/statements-remarks">Remarks and Statements</a>
      <a href="/news/press-releases/readouts">Readouts</a>
      <a href="/news/press-releases/testimonies">Testimonies</a>
      <a href="/news/press-releases/statements-remarks/secretary">Secretary Statements &amp; Remarks</a>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert [item.url for item in items] == ["https://home.treasury.gov/news/press-releases/sb0429"]


def test_aaa_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = AAAWatcher("aaa_test", "https://gasprices.aaa.com/")
    html = """
    <html><body>
      <a href="/national-gas-average-jumps-one-dollar-in-one-month/">National Gas Average Jumps One Dollar in One Month</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "AAA said national gas prices jumped sharply over the past month as crude and refining costs rose.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "AAA said national gas prices jumped sharply over the past month as crude and refining costs rose."



def test_conference_board_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = ConferenceBoardWatcher("conference_board_test", "https://www.conference-board.org/press/")
    html = """
    <html><body>
      <a href="/press/where-to-hire-index">The Best and Worst Places to Hire in Europe January 12, 2026</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "The Conference Board highlighted diverging hiring conditions across major European labor markets.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "The Conference Board highlighted diverging hiring conditions across major European labor markets."



def test_treasury_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = TreasuryWatcher("treasury_test", "https://home.treasury.gov/news/press-releases")
    html = """
    <html><body>
      <a href="/news/press-releases/sb0429">Secretary Bessent: Financial Literacy Unlocks Opportunity for Every American</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "Treasury said the initiative aims to expand financial literacy access and long-term household resilience.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "Treasury said the initiative aims to expand financial literacy access and long-term household resilience."


def test_census_retail_watcher_parses_official_page_and_workbook(monkeypatch):
    watcher = CensusRetailWatcher("census_retail_test", "https://www.census.gov/retail/sales.html")
    html = """
    <html><body>
      <div>FOR IMMEDIATE RELEASE: Wednesday, April 01, 2026</div>
      <h2>Advance Monthly Sales for Retail and Food Services</h2>
      <p>Advance estimates of U.S. retail and food services sales for February 2026, adjusted for seasonal variation, were $738.4 billion, up 0.6 percent from the previous month.</p>
      <a href="marts/www/marts_current.pdf">Advance Monthly Retail Trade Report</a>
      <a href="marts/www/marts_current.xlsx">Advance Monthly Retail Trade Report Tables</a>
    </body></html>
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1."
    ws.cell(row=6, column=10, value="Adjusted2")
    ws.cell(row=12, column=2, value="Retail & food services, total")
    ws.cell(row=12, column=10, value=738366)
    ws.cell(row=12, column=11, value=733955)
    ws.cell(row=13, column=2, value="Total (excl. motor vehicle & parts)")
    ws.cell(row=13, column=10, value=598988)
    ws.cell(row=13, column=11, value=596224)
    buffer = BytesIO()
    wb.save(buffer)

    monkeypatch.setattr(watcher, "_fetch_page_html", lambda: html)
    monkeypatch.setattr(watcher, "_fetch_binary", lambda url: buffer.getvalue())

    items = watcher._fetch_events()

    assert len(items) == 1
    assert items[0].title == "Advance Monthly Sales for Retail and Food Services: 0.6% m/m, ex-auto 0.5% m/m (February 2026)"
    assert items[0].url == "https://www.census.gov/retail/marts/www/marts_current.pdf"
    assert items[0].raw["total_pct"] == 0.6
    assert items[0].raw["ex_auto_pct"] == 0.5


def test_axios_watcher_deduplicates_duplicate_sitemap_urls(monkeypatch):
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    duplicate_entries = [
        SitemapEntry(url="https://www.axios.com/2026/03/31/china-pakistan-iran-peace-deal-strait-ceasefire", lastmod="2026-03-31T12:00:00Z"),
        SitemapEntry(url="https://www.axios.com/2026/03/31/china-pakistan-iran-peace-deal-strait-ceasefire", lastmod="2026-03-31T12:00:00Z"),
        SitemapEntry(url="https://www.axios.com/2026/03/31/iran-fbi-leaks-lockheed-martin-cyber-warfare", lastmod="2026-03-31T11:00:00Z"),
    ]
    monkeypatch.setattr(watcher, "_iter_sitemap_entries", lambda: duplicate_entries)

    items = watcher._fetch_events()

    assert [item.url for item in items] == [
        "https://www.axios.com/2026/03/31/china-pakistan-iran-peace-deal-strait-ceasefire",
        "https://www.axios.com/2026/03/31/iran-fbi-leaks-lockheed-martin-cyber-warfare",
    ]


def test_nyt_watcher_filters_official_rss_entries(monkeypatch):
    watcher = NytWatcher("nyt_test", "https://www.nytimes.com/section/business")
    monkeypatch.setattr(
        watcher,
        "_fetch_feed_entries",
        lambda feed_url: [
            NytFeedEntry(
                url="https://www.nytimes.com/2026/04/03/business/jobs-report-march.html",
                title="Job Growth Rebounds as Hiring Accelerates in March",
                summary="The jobs report showed 178,000 payroll gains and firmer wage pressure.",
                published_at="2026-04-03T12:30:00+00:00",
                feed_url=feed_url,
            ),
            NytFeedEntry(
                url="https://www.nytimes.com/2026/04/03/arts/design/museum-feature.html",
                title="A New Museum Wing Opens in Paris",
                summary="The renovation expands the museum's modern art collection.",
                published_at="2026-04-03T11:00:00+00:00",
                feed_url=feed_url,
            ),
            NytFeedEntry(
                url="https://www.nytimes.com/live/2026/04/03/world/iran-war-trump-oil/heres-the-latest",
                title="Here’s the latest.",
                summary="",
                published_at="2026-04-03T12:45:00+00:00",
                feed_url=feed_url,
            ),
        ],
    )

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["Job Growth Rebounds as Hiring Accelerates in March"]


def test_wsj_watcher_filters_official_rss_entries(monkeypatch):
    watcher = WsjWatcher("wsj_test", "https://www.wsj.com/economy")
    monkeypatch.setattr(
        watcher,
        "_fetch_feed_entries",
        lambda feed_url: [
            WsjFeedEntry(
                url="https://www.wsj.com/economy/jobs/private-sector-job-growth-steady-in-march-per-adp-9dab5be9",
                title="Private-Sector Job Growth Steady in March",
                summary="ADP said private payrolls rose in March.",
                published_at="2026-04-01T00:00:00+00:00",
                feed_url=feed_url,
            ),
            WsjFeedEntry(
                url="https://www.wsj.com/articles/target-drops-dei-goals-and-ends-program-to-boost-black-suppliers-77cb4c75",
                title="Target Drops DEI Goals and Ends Program to Boost Black Suppliers",
                summary="Retailer reshapes internal program.",
                published_at="2026-04-01T00:00:00+00:00",
                feed_url=feed_url,
            ),
        ],
    )

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["Private-Sector Job Growth Steady in March"]


def test_axios_watcher_prefers_official_news_sitemap_title_over_slug(monkeypatch):
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    monkeypatch.setattr(
        watcher,
        "_iter_sitemap_entries",
        lambda: [
            SitemapEntry(
                url="https://www.axios.com/2026/04/01/trump-to-address-nation-on-iran-wednesday-after-hinting-at-wars-end",
                lastmod="2026-04-01T00:41:52Z",
                title='Trump to address nation on Iran war Wednesday after saying US will leave "soon"',
                keywords="Iran,Trump,White House",
            )
        ],
    )

    items = watcher._fetch_events()

    assert [item.title for item in items] == ['Trump to address nation on Iran war Wednesday after saying US will leave "soon"']


def test_axios_watcher_prefers_richer_duplicate_entry_metadata(monkeypatch):
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    monkeypatch.setattr(
        watcher,
        "_iter_sitemap_entries",
        lambda: [
            SitemapEntry(
                url="https://www.axios.com/2026/03/31/china-pakistan-iran-peace-deal-strait-ceasefire",
                lastmod="2026-03-31T11:00:00Z",
            ),
            SitemapEntry(
                url="https://www.axios.com/2026/03/31/china-pakistan-iran-peace-deal-strait-ceasefire",
                lastmod="2026-03-31T11:00:00Z",
                title="China, Pakistan push Iran ceasefire plan",
                keywords="China,Pakistan,Iran",
            ),
        ],
    )

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["China, Pakistan push Iran ceasefire plan"]


def test_axios_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    monkeypatch.setattr(
        watcher,
        "_iter_sitemap_entries",
        lambda: [
            SitemapEntry(
                url="https://www.axios.com/2026/04/01/trump-to-address-nation-on-iran-wednesday-after-hinting-at-wars-end",
                lastmod="2026-04-01T00:41:52Z",
                title="Trump to address nation on Iran war Wednesday",
                keywords="Iran,Trump,White House",
            )
        ],
    )
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "Axios reported Trump planned a national address as ceasefire signaling and Iran headlines intensified.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "Axios reported Trump planned a national address as ceasefire signaling and Iran headlines intensified."


def test_axios_watcher_orders_matches_by_recency_and_keeps_primary_keyword_diagnostics(monkeypatch):
    monkeypatch.setenv("AXIOS_KEYWORDS", "iran,trump")
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    monkeypatch.setattr(
        watcher,
        "_iter_sitemap_entries",
        lambda: [
            SitemapEntry(
                url="https://www.axios.com/2026/04/02/trump-dhs-worker-pay-executive-order-congress-shutdown",
                lastmod="2026-04-02T15:00:00Z",
                title="Trump to circumvent Congress with order to pay all DHS workers",
                keywords="Donald Trump,Government shutdown,Top Stories",
            ),
            SitemapEntry(
                url="https://www.axios.com/2026/04/02/trump-bomb-iran-stone-ages-power-plants",
                lastmod="2026-04-02T14:00:00Z",
                title='Trump: US to bomb Iran "back to stone ages" over next 2-3 weeks',
                keywords="Iran deal,Iran,Nuclear program of Iran,Top Stories",
            ),
        ],
    )

    items = watcher._fetch_events()

    assert items[0].url == "https://www.axios.com/2026/04/02/trump-dhs-worker-pay-executive-order-congress-shutdown"
    assert items[1].url == "https://www.axios.com/2026/04/02/trump-bomb-iran-stone-ages-power-plants"
    assert items[1].raw["primary_matched_keywords"] == ["iran"]
    assert items[1].raw["primary_title_url_matched_keywords"] == ["iran"]
    assert items[1].raw["title_url_matched_keywords"] == ["iran", "trump"]
    assert items[1].raw["primary_title_url_match_count"] > items[0].raw["primary_title_url_match_count"]


def test_axios_blocked_article_fetch_sets_backoff_and_uses_keyword_fallback(monkeypatch):
    watcher = AxiosWatcher("axios_test", "https://www.axios.com/world/iran")
    url_one = "https://www.axios.com/2026/04/02/trump-bomb-iran-stone-ages-power-plants"
    url_two = "https://www.axios.com/2026/04/02/iran-foreign-minister-hormuz-shipping-oman"

    response = requests.Response()
    response.status_code = 403
    response._content = b"Forbidden"
    response.url = url_one

    calls = {"count": 0}

    def fake_fetch(session, url, accept=None):
        calls["count"] += 1
        raise requests.HTTPError(response=response)

    monkeypatch.setattr(axios_module, "fetch_with_retries", fake_fetch)
    monkeypatch.setattr(
        axios_module,
        "browser_impersonated_get",
        lambda *args, **kwargs: SimpleNamespace(status_code=403, text="Forbidden"),
    )

    assert watcher._fetch_article_summary(url_one) == ""
    assert watcher.summary_fetch_backoff_until > time.time()
    assert watcher._fetch_article_summary(url_two) == ""
    assert calls["count"] == 1

    item = Event(
        source=watcher.source_name,
        item_id="blocked-axios-item",
        title='Trump: US to bomb Iran "back to stone ages" over next 2-3 weeks',
        url=url_one,
        raw={"keywords": "Iran,Nuclear program of Iran,Top Stories"},
    )
    watcher._enrich_summaries([item])
    assert item.summary == 'Trump: US to bomb Iran "back to stone ages" over next 2-3 weeks'


def test_bloomberg_watcher_filters_official_news_sitemap_entries(monkeypatch):
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-03-31/fed-official-says-inflation-risks-remain</loc>
        <news:news>
          <news:publication_date>2026-03-31T19:00:00Z</news:publication_date>
          <news:title>Fed Official Says Inflation Risks Remain</news:title>
        </news:news>
      </url>
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-03-31/nike-slump-deepens-after-pessimistic-outlook</loc>
        <news:news>
          <news:publication_date>2026-03-31T18:00:00Z</news:publication_date>
          <news:title>Nike Slump Deepens After Pessimistic Outlook</news:title>
        </news:news>
      </url>
      <url>
        <loc>https://www.bloomberg.com/opinion/articles/2026-03-31/how-to-ride-out-market-panic</loc>
        <news:news>
          <news:publication_date>2026-03-31T17:00:00Z</news:publication_date>
          <news:title>How to Ride Out Market Panic</news:title>
        </news:news>
      </url>
    </urlset>
    """

    monkeypatch.setattr(watcher, "_fetch_xml_root", lambda: ET.fromstring(xml.encode("utf-8")))

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["Fed Official Says Inflation Risks Remain"]


def test_bloomberg_keywords_default_to_trade_symbols(monkeypatch):
    monkeypatch.delenv("BLOOMBERG_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BTC-USDC:BTC,BRENTOIL-USDC:xyz:BRENTOIL,SILVER-USDC:xyz:SILVER")

    keywords = _keywords_for_target("https://www.bloomberg.com/markets/economics")

    assert "bitcoin" in keywords
    assert "oil" in keywords
    assert "silver" in keywords
    assert "opec" in keywords
    assert "inflation" in keywords


def test_bloomberg_trade_symbol_profiles_include_broader_iran_updates_for_brent(monkeypatch):
    monkeypatch.delenv("BLOOMBERG_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.bloomberg.com/news/newsletters/2026-03-31/nato-allies-balk-at-us-call-for-help-in-war-with-iran-evening-briefing-americas</loc>
        <news:news>
          <news:publication_date>2026-03-31T22:36:07.948Z</news:publication_date>
          <news:title>NATO Allies Balk at US Call for Help in War With Iran: Evening Briefing Americas</news:title>
        </news:news>
      </url>
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-03-31/oil-prices-fall-sharply-wti-trades-near-102-a-barrel</loc>
        <news:news>
          <news:publication_date>2026-03-31T21:00:00Z</news:publication_date>
          <news:title>Oil Prices Drop as Iran Signals Readiness to End War</news:title>
        </news:news>
      </url>
    </urlset>
    """

    monkeypatch.setattr(watcher, "_fetch_xml_root", lambda: ET.fromstring(xml.encode("utf-8")))

    items = watcher._fetch_events()

    assert [item.title for item in items] == [
        "NATO Allies Balk at US Call for Help in War With Iran: Evening Briefing Americas",
        "Oil Prices Drop as Iran Signals Readiness to End War",
    ]


def test_bloomberg_watcher_accepts_live_blog_iran_updates_for_brent(monkeypatch):
    monkeypatch.delenv("BLOOMBERG_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.bloomberg.com/news/live-blog/2026-03-31/iran-war-updates-trump-markets-optimism</loc>
        <news:news>
          <news:publication_date>2026-03-31T21:00:00Z</news:publication_date>
          <news:title>Iran War Updates: Trump, Markets, Optimism</news:title>
        </news:news>
      </url>
    </urlset>
    """

    monkeypatch.setattr(watcher, "_fetch_xml_root", lambda: ET.fromstring(xml.encode("utf-8")))

    items = watcher._fetch_events()

    assert [item.url for item in items] == [
        "https://www.bloomberg.com/news/live-blog/2026-03-31/iran-war-updates-trump-markets-optimism"
    ]


def test_bloomberg_explicit_keywords_are_additive_to_trade_profiles(monkeypatch):
    monkeypatch.setenv("BLOOMBERG_KEYWORDS", "trump")
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-04-01/trump-tells-aides-hes-willing-to-end-war-without-reopening-hormuz</loc>
        <news:news>
          <news:publication_date>2026-04-01T18:00:00Z</news:publication_date>
          <news:title>Trump Tells Aides He’s Willing to End War Without Reopening Hormuz</news:title>
        </news:news>
      </url>
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-04-01/oil-jumps-after-middle-east-shipping-risks-rise</loc>
        <news:news>
          <news:publication_date>2026-04-01T17:00:00Z</news:publication_date>
          <news:title>Oil Jumps After Middle East Shipping Risks Rise</news:title>
        </news:news>
      </url>
    </urlset>
    """

    monkeypatch.setattr(watcher, "_fetch_xml_root", lambda: ET.fromstring(xml.encode("utf-8")))

    items = watcher._fetch_events()

    assert [item.title for item in items] == [
        "Trump Tells Aides He’s Willing to End War Without Reopening Hormuz",
        "Oil Jumps After Middle East Shipping Risks Rise",
    ]


def test_bloomberg_always_includes_dxy_and_treasury_topics(monkeypatch):
    monkeypatch.delenv("BLOOMBERG_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "SILVER-USDC:xyz:SILVER")

    keywords = _keywords_for_target("https://www.bloomberg.com/markets/economics")

    assert "dxy" in keywords
    assert "dollar index" in keywords
    assert "treasury" in keywords
    assert "treasury yields" in keywords


def test_coindesk_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = CoinDeskWatcher("coindesk_test", "https://www.coindesk.com/latest-crypto-news")
    html = """
    <html><body>
      <a href="/markets/2026/04/02/bitcoin-rallies-as-risk-appetite-returns/">Bitcoin Rallies as Risk Appetite Returns</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_listing_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "CoinDesk said bitcoin rose as macro risk sentiment improved and traders rotated back into crypto.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "CoinDesk said bitcoin rose as macro risk sentiment improved and traders rotated back into crypto."



def test_dol_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = DolWatcher("dol_test", "https://www.dol.gov/newsroom/releases")
    html = """
    <html><body>
      <a href="/newsroom/releases/eta/eta20260402">Labor Department Announces Workforce Development Initiative</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "The Labor Department said the initiative expands apprenticeship and workforce training access nationwide.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "The Labor Department said the initiative expands apprenticeship and workforce training access nationwide."


def test_bloomberg_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    monkeypatch.delenv("BLOOMBERG_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.bloomberg.com/news/articles/2026-04-01/oil-jumps-after-middle-east-shipping-risks-rise</loc>
        <news:news>
          <news:publication_date>2026-04-01T17:00:00Z</news:publication_date>
          <news:title>Oil Jumps After Middle East Shipping Risks Rise</news:title>
        </news:news>
      </url>
    </urlset>
    """

    monkeypatch.setattr(watcher, "_fetch_xml_root", lambda: ET.fromstring(xml.encode("utf-8")))
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "Oil rose as Middle East shipping risks climbed and traders repriced supply disruption odds.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "Oil Jumps After Middle East Shipping Risks Rise"
    assert items[0].summary == "Oil rose as Middle East shipping risks climbed and traders repriced supply disruption odds."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_bloomberg_blocked_article_fetch_sets_backoff_and_uses_title_fallback(monkeypatch):
    watcher = BloombergWatcher("bloomberg_test", "https://www.bloomberg.com/markets/economics")
    url_one = "https://www.bloomberg.com/news/articles/2026-04-01/oil-jumps-after-middle-east-shipping-risks-rise"
    url_two = "https://www.bloomberg.com/news/articles/2026-04-02/another-oil-update"
    calls = {"count": 0}

    def fake_fetch(_session, _url, *, accept=None):
        calls["count"] += 1
        response = requests.Response()
        response.status_code = 403
        response._content = b"forbidden"
        response.url = _url
        raise requests.HTTPError(response=response)

    monkeypatch.setattr(bloomberg_module, "fetch_with_retries", fake_fetch)
    monkeypatch.setattr(
        bloomberg_module,
        "browser_impersonated_get",
        lambda *args, **kwargs: SimpleNamespace(status_code=403, text="forbidden"),
    )

    assert watcher._fetch_article_summary(url_one) == ""
    assert watcher.summary_fetch_backoff_until > time.time()
    assert watcher._fetch_article_summary(url_two) == ""
    assert calls["count"] == 1

    item = Event(
        source=watcher.source_name,
        item_id="blocked-bloomberg-item",
        title="Oil Jumps After Middle East Shipping Risks Rise",
        url=url_one,
    )
    watcher._enrich_summaries([item])
    assert item.summary == "Oil Jumps After Middle East Shipping Risks Rise"


def test_mni_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = MniWatcher("mni_test", "https://www.mnimarkets.com/")
    html = """
    <html><body>
      <a href="/articles/headlines-from-trumps-iran-speech-crossing-1771496194958">Headlines From Trump's Iran Speech Crossing</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "Trump's speech headlines highlighted Iran, Hormuz passage restrictions, and broader political risk.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "Headlines From Trump's Iran Speech Crossing"
    assert items[0].summary == "Trump's speech headlines highlighted Iran, Hormuz passage restrictions, and broader political risk."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_white_house_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = WhiteHouseWatcher("white_house_test", "https://www.whitehouse.gov/news/")
    html = """
    <html><body>
      <a href="/remarks/2026/04/president-trump-delivers-powerful-primetime-address-on-operation-epic-fury/">President Trump Delivers Powerful Primetime Address on Operation Epic Fury</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "President Trump described the operation and next steps for restoring deterrence.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "President Trump Delivers Powerful Primetime Address on Operation Epic Fury"
    assert items[0].summary == "President Trump described the operation and next steps for restoring deterrence."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_reuters_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    monkeypatch.delenv("REUTERS_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/world/middle-east/oil-jumps-as-opec-output-falls-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T00:47:43.825Z</news:publication_date>
          <news:title>Oil jumps as OPEC output falls and traders watch Hormuz</news:title>
          <news:keywords>oil,opec,shipping</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "Oil jumped after OPEC output fell and traders watched Hormuz shipping risks.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "Oil jumps as OPEC output falls and traders watch Hormuz"
    assert items[0].summary == "Oil jumped after OPEC output fell and traders watched Hormuz shipping risks."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_reuters_blocked_article_fetch_sets_backoff_and_uses_keyword_fallback(monkeypatch):
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    url_one = "https://www.reuters.com/world/middle-east/oil-jumps-as-opec-output-falls-2026-04-01/"
    url_two = "https://www.reuters.com/world/middle-east/another-oil-update-2026-04-02/"
    calls = {"count": 0}

    def fake_fetch(_session, _url, *, accept=None):
        calls["count"] += 1
        response = requests.Response()
        response.status_code = 401
        response._content = b"access denied"
        response.url = _url
        raise requests.HTTPError(response=response)

    monkeypatch.setattr(reuters_module, "fetch_with_retries", fake_fetch)
    monkeypatch.setattr(
        reuters_module,
        "browser_impersonated_get",
        lambda *args, **kwargs: SimpleNamespace(status_code=401, text="access denied"),
    )

    assert watcher._fetch_article_summary(url_one) == ""
    assert watcher.summary_fetch_backoff_until > time.time()
    assert watcher._fetch_article_summary(url_two) == ""
    assert calls["count"] == 1

    item = Event(
        source=watcher.source_name,
        item_id="blocked-reuters-item",
        title="Oil jumps as OPEC output falls and traders watch Hormuz",
        url=url_one,
        raw={"keywords": "oil,opec,shipping"},
    )
    watcher._enrich_summaries([item])
    assert item.summary == "Oil jumps as OPEC output falls and traders watch Hormuz. Topics: oil, opec, shipping."


def test_reuters_trade_profiles_match_adp_jobs_title(monkeypatch):
    monkeypatch.delenv("REUTERS_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BTC-USDC:BTC")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/world/us/private-sector-job-growth-steady-march-adp-says-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T12:00:00Z</news:publication_date>
          <news:title>Private-Sector Job Growth Steady in March, ADP Says</news:title>
          <news:keywords>jobs,ADP,employment</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["Private-Sector Job Growth Steady in March, ADP Says"]


def test_reuters_trade_profiles_match_iran_ceasefire_update(monkeypatch):
    monkeypatch.delenv("REUTERS_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/world/middle-east/iran-says-trumps-statements-tehran-requesting-ceasefire-are-false-baseless-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T13:00:00Z</news:publication_date>
          <news:title>Iran says Trump's statements on Tehran requesting ceasefire are false, baseless</news:title>
          <news:keywords>Iran,ceasefire,Middle East</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)

    items = watcher._fetch_events()

    assert [item.url for item in items] == [
        "https://www.reuters.com/world/middle-east/iran-says-trumps-statements-tehran-requesting-ceasefire-are-false-baseless-2026-04-01/"
    ]


def test_reuters_watcher_excludes_localized_non_english_paths(monkeypatch):
    monkeypatch.delenv("REUTERS_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/pt/mundo/MGKW3TNMSROMFFQABCVCGT3NZA-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T18:56:26.503Z</news:publication_date>
          <news:title>Presidente iraniano diz em carta que Irã não tem inimizade com norte-americanos comuns</news:title>
          <news:keywords>Iran,ceasefire,Middle East</news:keywords>
        </news:news>
      </url>
      <url>
        <loc>https://www.reuters.com/world/middle-east/iran-says-trumps-statements-tehran-requesting-ceasefire-are-false-baseless-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T13:00:00Z</news:publication_date>
          <news:title>Iran says Trump's statements on Tehran requesting ceasefire are false, baseless</news:title>
          <news:keywords>Iran,ceasefire,Middle East</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)

    items = watcher._fetch_events()

    assert [item.url for item in items] == [
        "https://www.reuters.com/world/middle-east/iran-says-trumps-statements-tehran-requesting-ceasefire-are-false-baseless-2026-04-01/"
    ]


def test_factbase_watcher_parses_split_date_lines_and_full_text_marker():
    watcher = FactbaseWatcher("factbase_test", "https://rollcall.com/factbase/trump/topic/calendar/")
    html = """
    <html><body>
      <div>President's Public Schedule</div>
      <div>Tuesday,</div>
      <div>March 31 2026</div>
      <div>1:00 PM</div>
      <div>1:00 PM</div>
      <div>Press Briefing by the White House Press Secretary Karoline Leavitt</div>
      <div>Full Text and Analysis</div>
      <div>James S. Brady Press Briefing Room</div>
      <div>On Camera</div>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert len(items) == 1
    assert items[0].title == "Press Briefing by the White House Press Secretary Karoline Leavitt"
    assert "Tuesday, March 31 2026" in items[0].summary
    assert "1:00 PM" in items[0].summary
    assert "James S. Brady Press Briefing Room" in items[0].summary


def test_dol_watcher_raises_helpful_error_on_challenge(monkeypatch):
    watcher = DolWatcher("dol_test", "https://www.dol.gov/newsroom/releases")
    challenge_html = "<html><title>Challenge Validation</title><script>function cp_clge_done(){}</script><iframe class='custmsg crypto'></iframe></html>"
    monkeypatch.setattr(watcher, "_fetch_html_with_browser_impersonation", lambda: challenge_html)

    with pytest.raises(RuntimeError, match="DOL returned a challenge page"):
        watcher._fetch_html()


def test_bls_watcher_raises_helpful_error_on_access_denied(monkeypatch):
    watcher = BlsWatcher("bls_test", "https://www.bls.gov/feed/bls_latest.rss")

    class FakeResponse:
        status_code = 403
        text = "<html><title>Access Denied</title><h1>Bureau of Labor Statistics</h1></html>"
        headers = {"Content-Type": "text/html"}

    monkeypatch.setattr(watcher, "_fetch_feed_response", lambda: FakeResponse())

    with pytest.raises(RuntimeError, match="BLS denied access"):
        watcher._fetch_events()


def test_bls_watcher_defaults_to_web_poll_interval():
    watcher = BlsWatcher("bls_test", "https://www.bls.gov/feed/bls_latest.rss")

    assert watcher.interval_seconds == WEB_POLL_SECONDS


def test_irna_watcher_parses_telegram_fallback_posts():
    watcher = IrnaWatcher("irna_test", "https://en.irna.ir/")
    html = """
    <html><body>
      <div class="tgme_widget_message_wrap">
        <div class="tgme_widget_message" data-post="Irna_en/33567">
          <a class="tgme_widget_message_date" href="https://t.me/Irna_en/33567">
            <time class="time" datetime="2026-03-30T17:17:49+00:00">17:17</time>
          </a>
          <div class="tgme_widget_message_text" dir="auto">
            <i>🔸</i>
            <b>Provocative acts in Strait of Hormuz to further exacerbate the situation: Araghchi</b>
            <br/>
            <blockquote>Tehran, IRNA – Foreign Minister Abbas Araghchi says any provocative action would further exacerbate the situation.</blockquote>
            <br/>
            <a href="https://t.me/Irna_en" target="_blank">@Irna_en</a>
          </div>
        </div>
      </div>
    </body></html>
    """

    items = watcher.extract_events(html)

    assert len(items) == 1
    assert items[0].url == "https://t.me/Irna_en/33567"
    assert items[0].published_at == "2026-03-30T17:17:49+00:00"
    assert items[0].title == "Provocative acts in Strait of Hormuz to further exacerbate the situation: Araghchi"
    assert "Foreign Minister Abbas Araghchi" in items[0].summary
    assert items[0].raw["fallback"] == "telegram_public_page"


def test_irna_watcher_falls_back_to_telegram_when_official_site_challenges(monkeypatch):
    watcher = IrnaWatcher("irna_test", "https://en.irna.ir/")
    challenge_html = "<html><body>Transferring to the website...<script>__arcsjs=1</script><div>__arcsjsc</div></body></html>"
    telegram_html = """
    <html><body>
      <div class="tgme_widget_message_wrap">
        <div class="tgme_widget_message">
          <a class="tgme_widget_message_date" href="https://t.me/Irna_en/33567"><time datetime="2026-03-30T17:17:49+00:00">17:17</time></a>
          <div class="tgme_widget_message_text"><b>Fallback works</b></div>
        </div>
      </div>
    </body></html>
    """

    monkeypatch.setattr(HtmlPageWatcher, "_fetch_html", lambda self: challenge_html)
    monkeypatch.setattr(watcher, "_fetch_telegram_html", lambda: telegram_html)

    html = watcher._fetch_html()

    assert html == telegram_html


def test_reuters_watcher_filters_official_news_sitemap_entries(monkeypatch):
    monkeypatch.delenv("REUTERS_KEYWORDS", raising=False)
    monkeypatch.setenv("TRADE_SYMBOLS", "BTC-USDC:BTC,BRENTOIL-USDC:xyz:BRENTOIL,SILVER-USDC:xyz:SILVER")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/world/middle-east/oil-jumps-as-opec-output-falls-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T00:47:43.825Z</news:publication_date>
          <news:title>Oil jumps as OPEC output falls and traders watch Hormuz</news:title>
          <news:keywords>oil,opec,shipping</news:keywords>
        </news:news>
      </url>
      <url>
        <loc>https://www.reuters.com/world/us/nike-falls-after-outlook-cut-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T00:40:00.000Z</news:publication_date>
          <news:title>Nike falls after company cuts annual outlook</news:title>
          <news:keywords>retail,stocks</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)

    items = watcher._fetch_events()

    assert [item.title for item in items] == ["Oil jumps as OPEC output falls and traders watch Hormuz"]


def test_reuters_watcher_explicit_keywords_are_additive_to_trade_profiles(monkeypatch):
    monkeypatch.setenv("REUTERS_KEYWORDS", "payroll,employment")
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL-USDC:xyz:BRENTOIL")
    watcher = ReutersWatcher("reuters_test", "https://www.reuters.com/markets/")
    index_xml = """
    <sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
      <sitemap><loc>https://www.reuters.com/arc/outboundfeeds/news-sitemap/?outputType=xml</loc></sitemap>
    </sitemapindex>
    """
    page_xml = """
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">
      <url>
        <loc>https://www.reuters.com/world/us/us-payrolls-rise-more-than-expected-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T12:00:00.000Z</news:publication_date>
          <news:title>US payrolls rise more than expected in March</news:title>
          <news:keywords>jobs,payroll,employment</news:keywords>
        </news:news>
      </url>
      <url>
        <loc>https://www.reuters.com/world/middle-east/oil-jumps-as-opec-output-falls-2026-04-01/</loc>
        <news:news>
          <news:publication_date>2026-04-01T11:00:00.000Z</news:publication_date>
          <news:title>Oil jumps as OPEC output falls</news:title>
          <news:keywords>oil,opec</news:keywords>
        </news:news>
      </url>
    </urlset>
    """

    def fake_fetch(url: str):
        if "news-sitemap-index" in url:
            return ET.fromstring(index_xml.encode("utf-8"))
        if "news-sitemap/?" in url:
            return ET.fromstring(page_xml.encode("utf-8"))
        raise AssertionError(url)

    monkeypatch.setattr(watcher, "_fetch_xml_root", fake_fetch)

    items = watcher._fetch_events()

    assert [item.title for item in items] == [
        "US payrolls rise more than expected in March",
        "Oil jumps as OPEC output falls",
    ]


def test_schwab_watcher_poll_enriches_missing_summary_from_article(monkeypatch, tmp_path):
    watcher = SchwabWatcher("schwab_test", "https://www.schwab.com/learn/market-commentary")
    html = """
    <html><body>
      <a href="/learn/story/stock-market-update-open">
        <h3>Job Creation Rebounds in March, Unemployment Dips</h3>
        <span>Article | Apr 3, 2026</span>
      </a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "The U.S. reported 178,000 new jobs in March and a 4.3% unemployment rate.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "Job Creation Rebounds in March, Unemployment Dips"
    assert items[0].summary == "The U.S. reported 178,000 new jobs in March and a 4.3% unemployment rate."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_spglobal_pmi_watcher_poll_enriches_new_items_with_article_summary(monkeypatch, tmp_path):
    watcher = SpglobalPmiWatcher("spglobal_pmi_test", "https://www.spglobal.com/marketintelligence/en/mi/research-analysis/pmi.html")
    html = """
    <html><body>
      <a href="/marketintelligence/en/mi/research-analysis/monthly-pmi-bulletin-february-2026.html">Monthly PMI Bulletin: February 2026</a>
    </body></html>
    """
    monkeypatch.setattr(watcher, "_fetch_html", lambda: html)
    monkeypatch.setattr(
        watcher,
        "_fetch_article_summary",
        lambda url: "The global economic expansion picked up pace at the start of 2026 amid improvements across manufacturing and services.",
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].title == "Monthly PMI Bulletin: February 2026"
    assert items[0].summary == "The global economic expansion picked up pace at the start of 2026 amid improvements across manufacturing and services."
    assert state.is_seen(watcher.source_name, items[0].item_id)


def test_nyt_watcher_poll_preserves_feed_summary(monkeypatch, tmp_path):
    monkeypatch.setenv("TRADE_SYMBOLS", "BRENTOIL")
    watcher = NytWatcher("nyt_test", "https://www.nytimes.com/section/business")
    monkeypatch.setattr(
        watcher,
        "_fetch_feed_entries",
        lambda feed_url: [
            NytFeedEntry(
                url="https://www.nytimes.com/2026/04/03/world/middleeast/iran-oil-shipping.html",
                title="Oil Rises as Traders Reprice Strait of Hormuz Risk",
                summary="Analysts said shipping risk premiums climbed after new threats around the waterway.",
                published_at="2026-04-03T15:00:00+00:00",
                feed_url=feed_url,
            )
        ],
    )

    state = StateStore(tmp_path / "state.json")
    items = watcher.poll(state)

    assert len(items) == 1
    assert items[0].summary == "Analysts said shipping risk premiums climbed after new threats around the waterway."
    assert state.is_seen(watcher.source_name, items[0].item_id)
